
import openpyxl
from openpyxl import load_workbook

book = load_workbook('AUTOMATION/SPECIAL ORDER TEMPLATE.xlsx')
sheet = book['order']
cell = sheet.cell(row = 1, column=1)


print(cell.value)