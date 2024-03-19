import pandas as pd
from openpyxl import load_workbook

source_file = 'AUTOMATION/To_order.xlsx'

# List of source columns to copy and their corresponding destination columns
source_columns_dest_columns = {
    "DocNum": 1,
    "LineNum": 2,
    "CODE": 3,
    "QUANTITY": 5,
    "STORE CODE": 14,
}

# Load the destination Excel file and select the sheet
destination_file = "AUTOMATION/POR1 - Document_Lines1.xlsx"
destination_sheet_name = "Sheet1"
destination_wb = load_workbook(destination_file)
destination_ws = destination_wb[destination_sheet_name]

# Paste the column data into the destination sheet
start_row = 3  # Change this to the desired starting row
for source_column, dest_col_idx in source_columns_dest_columns.items():
    source_df = pd.read_excel(source_file)
    source_column_data = source_df[source_column]

    for idx, value in enumerate(source_column_data, start=start_row):
        destination_ws.cell(row=idx, column=dest_col_idx, value=value)

# Save the updated destination Excel file
destination_wb.save(destination_file)