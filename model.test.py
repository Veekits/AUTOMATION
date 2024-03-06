import os
import pandas as pd
from sklearn.preprocessing import LabelEncoder
import re
from openpyxl import load_workbook

base_dir = 'C:/Users/VMUKITA/OneDrive - Goodlife Pharmacy/Desktop/MY PROJECTS/Automation/AUTOMATION/OUTPUT/Excel Attachments'

def clean_quantity(quantity):
    # Remove non-numeric characters and specific strings
    cleaned_quantity = re.sub(r'[^0-9.]', '', str(quantity))
    return int(float(cleaned_quantity)) if cleaned_quantity else None

def extract_codes_and_quantities(directory):
    # Create a list to store the file paths
    file_paths = []

    # Create an empty DataFrame to store the results
    result_df = pd.DataFrame(columns=['Code', 'Quantity'])

    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.xlsx') or file.endswith('.xlsm'):
                file_path = os.path.join(root, file)
                file_paths.append(file_path)

    # Sort the file paths based on their index in the folder name
    file_paths.sort(key=lambda x: int(os.path.basename(os.path.dirname(x)).split('_')[0]))

    # Initialize a label encoder for categorical encoding
    label_encoder = LabelEncoder()

    # Flag to indicate if the first folder has been processed
    first_folder_processed = False

    for file_path in file_paths:
        try:
            # Read the Excel file
            df = pd.read_excel(file_path)

            # Identify columns with codes and quantities
            code_columns = [col for col in df.columns if 'code' in str(col).lower() or 'item no' in str(col).lower()]
            quantity_columns = [col for col in df.columns if 'quantity' in str(col).lower() or 'qty' in str(col).lower()]

            if code_columns and quantity_columns:
                # Extract data from identified columns
                codes_raw = df[code_columns].values.flatten()
                quantities_raw = df[quantity_columns].values.flatten()

                # Clean codes and quantities
                codes = [int(code) if pd.notna(code) else None for code in codes_raw]
                quantities = [clean_quantity(q) for q in quantities_raw]

                # Filter out entries where the cleaned quantity is not None
                non_empty_entries = pd.DataFrame({'Code': [code for code in codes if code is not None],
                                                  'Quantity': [quantity for quantity in quantities if quantity is not None]})

                #Clean folder_name
                folder_name = os.path.basename(os.path.dirname(file_path))
                folder_name = re.sub(r'^\d+_', '', folder_name)

                #Append results with folder_name to DataFrame
                non_empty_entries['Folder_Name'] = folder_name
                folder_names = []
                folder_names.append(folder_name)
                result_df = pd.concat([result_df, non_empty_entries], ignore_index=True)

                # Print the results
                print(f"Folder: {folder_name}, File: {os.path.basename(file_path)}")
                for code, quantity in zip(non_empty_entries['Code'], non_empty_entries['Quantity']):
                    print(f"Code: {code}, Quantity: {quantity}")

                # Set the flag to indicate the first folder has been processed
                first_folder_processed = True

        except Exception as e:
            print(f"Error processing {file_path}: {e}")

        # Break out of the loop after processing the first folder
        if first_folder_processed:
            break

    # Write the results to the Excel file
    excel_path = 'AUTOMATION/SPECIAL ORDER TEMPLATE.xlsx'
    wb = load_workbook(excel_path)
    ws = wb['order']

    # Get the last row in the worksheet
    last_row = ws.max_row

    # Write the 'Code' column to Column A and 'Quantity' column to Column C
    for i, (code, quantity, folder_name) in enumerate(result_df[['Code','Quantity', 'Folder_Name']].itertuples(index=False), start=1):
        ws.cell(row=last_row + i, column=1, value=code)
        ws.cell(row=last_row + i, column=3, value=quantity)
        ws.cell(row=last_row + i, column=9, value=folder_name)

    # Save the workbook
    wb.save(excel_path)

# Call the function with the base directory
extract_codes_and_quantities(base_dir)
