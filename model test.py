import os
import pandas as pd
from sklearn.preprocessing import LabelEncoder
import re
from openpyxl import load_workbook
from datetime import datetime

base_dir = 'C:/Users/VMUKITA/OneDrive - Goodlife Pharmacy/Desktop/MY PROJECTS/Automation/AUTOMATION/OUTPUT/Excel Attachments'

def clean_quantity(quantity):
    # Remove non-numeric characters and specific strings
    cleaned_quantity = re.sub(r'[^0-9.]', '', str(quantity))
    return int(float(cleaned_quantity)) if cleaned_quantity else None

def clear_order_sheet(excel_file):
    # Clear 'order' sheet from 2nd row downward
    wb = load_workbook(excel_file)
    ws = wb['order']
    ws.delete_rows(2, ws.max_row)
    wb.save(excel_file)
    wb.close()

def clear_to_order_sheet(excel_file):
    # Clear entire 'Sheet1' in 'To_order.xlsx'
    wb = load_workbook(excel_file)
    ws = wb['Sheet1']
    ws.delete_rows(1, ws.max_row)
    wb.save(excel_file)
    wb.close()

def clear_data_sheet(excel_file, sheet_name, start_row):
    # Clear data in a specific sheet from a starting row downward
    wb = load_workbook(excel_file)
    ws = wb[sheet_name]
    ws.delete_rows(start_row, ws.max_row)
    wb.save(excel_file)
    wb.close()

def extract_codes_and_quantities(directory):
    # Create a list to store the file paths
    file_paths = []

    # Create an empty DataFrame to store the results
    result_df = pd.DataFrame(columns=['Code', 'Quantity'])

    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.xlsx') or file.endswith('.xlsm'):
                file_path = os.path.join(root, file)
                file_paths.append(file_path)

    # Sort the file paths based on their index in the folder name
    file_paths.sort(key=lambda x: int(os.path.basename(os.path.dirname(x)).split('_')[0]))

    # Initialize a label encoder for categorical encoding
    label_encoder = LabelEncoder()

    # Flag to indicate if the first folder has been processed
    first_folder_processed = False

    for file_path in file_paths:
        try:
            # Read the Excel file
            df = pd.read_excel(file_path)

            # Identify columns with codes and quantities
            code_columns = [col for col in df.columns if 'code' in str(col).lower() or 'item no' in str(col).lower()]
            quantity_columns = [col for col in df.columns if 'quantity' in str(col).lower() or 'qty' in str(col).lower()]

            if code_columns and quantity_columns:
                # Extract data from identified columns
                codes_raw = df[code_columns].values.flatten()
                quantities_raw = df[quantity_columns].values.flatten()

                # Clean codes and quantities
                codes = [int(code) if pd.notna(code) else None for code in codes_raw]
                quantities = [clean_quantity(q) for q in quantities_raw]

                # Filter out entries where the cleaned quantity is not None
                non_empty_entries = pd.DataFrame({'Code': [code for code in codes if code is not None],
                                                  'Quantity': [quantity for quantity in quantities if quantity is not None]})

                #Clean folder_name
                folder_name = os.path.basename(os.path.dirname(file_path))
                folder_name = re.sub(r'^\d+_', '', folder_name)

                #Append results with folder_name to DataFrame
                non_empty_entries['Folder_Name'] = folder_name
                folder_names = []
                folder_names.append(folder_name)
                result_df = pd.concat([result_df, non_empty_entries], ignore_index=True)

                # Print the results
                print(f"Folder: {folder_name}, File: {os.path.basename(file_path)}")
                for code, quantity in zip(non_empty_entries['Code'], non_empty_entries['Quantity']):
                    print(f"Code: {code}, Quantity: {quantity}")

                # Set the flag to indicate the first folder has been processed
                first_folder_processed = True

        except Exception as e:
            print(f"Error processing {file_path}: {e}")

        # Break out of the loop after processing the first folder
        if first_folder_processed:
            break

    # Write the results to the Excel file
    excel_path = 'AUTOMATION/SPECIAL ORDER TEMPLATE.xlsx'
    wb = load_workbook(excel_path)
    ws = wb['order']

    # Get the last row in the worksheet
    last_row = ws.max_row

    # Write the 'Code' column to Column A and 'Quantity' column to Column C
    for i, (code, quantity, folder_name) in enumerate(result_df[['Code','Quantity', 'Folder_Name']].itertuples(index=False), start=1):
        ws.cell(row=last_row + i, column=1, value=code)
        ws.cell(row=last_row + i, column=2, value=quantity)
        ws.cell(row=last_row + i, column=3, value=folder_name)

    # Save the workbook
    wb.save(excel_path)

    file_name = 'AUTOMATION/SPECIAL ORDER TEMPLATE.xlsx'

    order = pd.read_excel(file_name, sheet_name = 'order')
    details = pd.read_excel(file_name, sheet_name = 'details')
    whse = pd.read_excel(file_name, sheet_name = 'whse')
    supp_code = pd.read_excel(file_name, sheet_name = 'supp_code')

    order = pd.merge(order, whse, on = 'STORE NAME')
    order = pd.merge(order, details, on = 'CODE')
    order = pd.merge(order, supp_code, on = 'SUPPLIER')

    order.sort_values(by='BP Code', ascending=True, inplace=True)
    order['DocNum'] = order.groupby('BP Code').ngroup() + 1
    order['LineNum'] = order.groupby('BP Code').cumcount()

    order['DocType'] = 'dDocument_Items'
        # Additional code for clearing sheets
    clear_order_sheet('AUTOMATION/SPECIAL ORDER TEMPLATE.xlsx')
    clear_to_order_sheet('AUTOMATION/To_order.xlsx')

    order['Comments'] = 'SPECIAL ORDER'
    order['DocDate'] = datetime.today().strftime('%Y%m%d')
    order['DocDate2'] = datetime.today().strftime('%Y%m%d')

    order.to_excel('AUTOMATION/To_order.xlsx')

    source_file = 'AUTOMATION/To_order.xlsx'

    def clear_data_from_destination_sheets(destination_file1, destination_file2):
        # Clear data from the third row downward in 'POR1 - Document_Lines1.xlsx'
        wb = load_workbook(destination_file1)
        ws = wb['Sheet1']
        ws.delete_rows(3, ws.max_row)
        wb.save(destination_file1)
        wb.close()

        # Clear data from the third row downward in 'OPOR - Documents.xlt.xlsx'
        wb = load_workbook(destination_file2)
        ws = wb['Sheet1']
        ws.delete_rows(3, ws.max_row)
        wb.save(destination_file2)
        wb.close()

    # Call the function to clear data from the destination sheets
    clear_data_from_destination_sheets('AUTOMATION/POR1 - Document_Lines1.xlsx', 'AUTOMATION/OPOR - Documents.xlt.xlsx') 

    # List of source columns to copy and their corresponding destination columns
    source_columns_dest_columns = {
         "DocNum": 1,
        "DocType": 3,
        "DocDate": 6,
        "DocDate2": 7,
        "BP Code": 8,
        "Comments": 18,
    }

    # Load the destination Excel file and select the sheet
    destination_file = "AUTOMATION/OPOR - Documents.xlt.xlsx"
    destination_sheet_name = "Sheet1"
    destination_wb = load_workbook(destination_file)
    destination_ws = destination_wb[destination_sheet_name]

    # Paste the column data into the destination sheet
    start_row = 3  # Change this to the desired starting row
    for source_column, dest_col_idx in source_columns_dest_columns.items():
        source_df = pd.read_excel(source_file)
        source_column_data = source_df[source_column]

        for idx, value in enumerate(source_column_data, start=start_row):
            destination_ws.cell(row=idx, column=dest_col_idx, value=value)

    # Save the updated destination Excel file
    destination_wb.save(destination_file)

    source_file = 'AUTOMATION/To_order.xlsx'

    # List of source columns to copy and their corresponding destination columns
    source_columns_dest_columns = {
        "DocNum": 1,
        "LineNum": 2,
        "CODE": 3,
        "QUANTITY": 5,
        "STORE CODE": 14,
    }

    # Load the destination Excel file and select the sheet
    destination_file = "AUTOMATION/POR1 - Document_Lines1.xlsx"
    destination_sheet_name = "Sheet1"
    destination_wb = load_workbook(destination_file)
    destination_ws = destination_wb[destination_sheet_name]

    # Paste the column data into the destination sheet
    start_row = 3  # Change this to the desired starting row
    for source_column, dest_col_idx in source_columns_dest_columns.items():
        source_df = pd.read_excel(source_file)
        source_column_data = source_df[source_column]

        for idx, value in enumerate(source_column_data, start=start_row):
            destination_ws.cell(row=idx, column=dest_col_idx, value=value)

    # Save the updated destination Excel file
    destination_wb.save(destination_file)

# Call the function with the base directory
extract_codes_and_quantities(base_dir)


