import os
from openpyxl import load_workbook
import re

# Directory containing the Excel files
base_dir = 'C:/Users/VMUKITA/OneDrive - Goodlife Pharmacy/Desktop/MY PROJECTS/Automation/OUTPUT/Excel Attachments'

def find_code_and_quantity(directory):
    # Create a list to store the file paths
    file_paths = []

    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith('.xlsx'):
                file_path = os.path.join(root, file)
                file_paths.append(file_path)

    # Sort the file paths based on their index in the folder name
    file_paths.sort(key=lambda x: int(os.path.basename(os.path.dirname(x)).split('_')[0]))

    for file_path in file_paths:
        try:
            # Load the workbook
            wb = load_workbook(file_path)
            for sheet_name in wb.sheetnames:
                # Check if the sheet contains columns named "code" and "quantity" in the specified range
                sheet = wb[sheet_name]
                header_row_values = list(sheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=3, values_only=True))[0]

                code_index = next((i for i, val in enumerate(header_row_values) if 'code' in str(val).lower()), None)
                quantity_index = next((i for i, val in enumerate(header_row_values) if 'quantity' in str(val).lower()), None)

                if code_index is not None or quantity_index is not None:
                    # Print values under the 'code' and 'quantity' columns
                    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
                        code = row[code_index].value if code_index is not None else None
                        quantity_raw = row[quantity_index].value if quantity_index is not None else None

                        # Use regex to extract only numeric values
                        quantity = int(re.sub(r'\D', '', str(quantity_raw))) if quantity_raw is not None else None

                        if code is not None or quantity is not None:
                            print(f"Folder: {os.path.basename(os.path.dirname(file_path))}, File: {os.path.basename(file_path)}, Code: {code if code is not None else '[Empty]'}, Quantity: {quantity if quantity is not None else '[Empty]'}")
                        else:
                            print(f"Folder: {os.path.basename(os.path.dirname(file_path))}, File: {os.path.basename(file_path)}, Code: [Empty], Quantity: [Empty]")
                    break

        except Exception as e:
            print(f"Error processing {file_path}: {e}")

# Call the function with the base directory
find_code_and_quantity(base_dir)

